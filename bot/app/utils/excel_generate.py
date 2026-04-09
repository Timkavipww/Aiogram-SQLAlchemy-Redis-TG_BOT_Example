import io
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from aiogram.types import BufferedInputFile

async def generate_excel_file(data, title="Данные"):
    # Создаем рабочую книгу
    wb = Workbook()
    ws = wb.active
    safe_title = re.sub(r'[\/\\\?\*\[\]:]', "_", title)
    ws.title = safe_title[:31]

    # Заголовки
    headers = list(data[0].keys()) if isinstance(data, list) and data else ["Ключ", "Значение"]
    ws.append(headers)

    # Данные
    if isinstance(data, list):
        for row in data:
            ws.append([row.get(h, "") for h in headers])
    else:  # если пришел объект
        for k, v in data.items():
            ws.append([k, v])

    # Автоширина колонок
    for i, col in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value or "")) for cell in col) + 2
        ws.column_dimensions[get_column_letter(i)].width = max_length

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return BufferedInputFile(
        file_stream.read(),
        filename=f"{safe_title}.xlsx"
    )